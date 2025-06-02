import sys
import traceback
import re
import pdfplumber
import os
from datetime import datetime
import pandas as pd
from PyQt6.QtWidgets import (
    QApplication,
    QMainWindow,
    QPushButton,
    QLabel,
    QFileDialog,
    QVBoxLayout,
    QWidget,
    QTableWidget,
    QTableWidgetItem,
    QTabWidget,
    QMessageBox,
    QDialog
)

from PyQt6 import QtCore
from docx import Document
from openpyxl import Workbook, load_workbook
from excel_writer import writeWeek, reorder_tabs, acknowledge
from ExcelReader import ExcelReader
from constructor import Student
from pdf_parser import extract_students_from_pdf
from excel_compare import compare_students, write_summary_sheet

#Main App Window - GUI Class
class TruancyApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Truancy Tracking System")
        self.setGeometry(100, 100, 1000, 700)

        #store students parsed in a list
        self.pdf_students = []
        #setup dataframes
        #previous spreadsheet data
        self.previous_data = pd.DataFrame()
        #load current working
        self.current_data = pd.DataFrame()
        self.loaded_workbook = None  #store the currently loaded Excel workbook
        self.loaded_file_path = ""   #path to workbook


        self.initUI()

    def initUI(self):
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)

        self.dashboard_tab = QWidget()
        self.reports_tab = QWidget()
        #tabs for the GUI
        self.tabs.addTab(self.dashboard_tab, "📊 Dashboard & File Management")
        self.tabs.addTab(self.reports_tab, "Reports")

        self.setup_dashboard()
        self.setup_reports()

    def setup_dashboard(self):
        layout = QVBoxLayout()

        # import PDF button
        self.import_pdf_button = QPushButton("📂 Select PDF Report")
        self.import_pdf_button.clicked.connect(self.ingest_pdf)
        layout.addWidget(self.import_pdf_button)

        # load student data (from excel) button
        self.load_spreadsheet_button = QPushButton("📂 Select Excel Workbook")
        self.load_spreadsheet_button.clicked.connect(self.load_spreadsheet)
        layout.addWidget(self.load_spreadsheet_button)

        # create new export file manually
        self.new_file_button = QPushButton("🆕 Create New Excel Workbook")
        self.new_file_button.clicked.connect(self.create_new_export_file)
        layout.addWidget(self.new_file_button)

        # export to Excel
        self.export_excel_button = QPushButton("🔄 Update Current Excel Workbook")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        self.export_excel_button.setEnabled(False)

        layout.addWidget(self.export_excel_button)

        self.warning_label = QLabel("")
        layout.addWidget(self.warning_label)

        # table to show PDF data
        self.pdf_table = QTableWidget()
        layout.addWidget(self.pdf_table)

        # show current active export file
        self.current_file_label = QLabel("📁 No export file loaded.")
        layout.addWidget(self.current_file_label)


        self.dashboard_tab.setLayout(layout)
    #Loads spreadsheet data and parse students
    def load_spreadsheet(self):
        #open file dialog to select spreadsheet
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Open Student Data",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv)"
        )
        if not file_path:
            return
        #store selected path and update UI
        self.loaded_file_path = file_path
        self.export_excel_button.setEnabled(True)
        self.current_file_label.setText(f"📁 Current File: {os.path.basename(file_path)}")
        
        #load workbook and extract data using Excel reader
        self.loaded_workbook = load_workbook(file_path)
        reader = ExcelReader(file_path)
        reader.read()
        #convert parsed student objects to Dataframe
        self.current_data = pd.DataFrame([s.__dict__ for s in reader.fixed])

        print("Loaded spreadsheet data:")
        print(self.current_data.head())

    #Will remove this method since we are displaying summary sheet in Excel
    def setup_reports(self):
        layout = QVBoxLayout()
        self.report_tabs = QTabWidget()

        # categories for the Reports tab (this may change or not be implemented at all)
        categories = [
            "Mediation Threshold",
            "Awaiting Mediation Letter",
            "Bounced Letters",
            "Students Awaiting Second Letter"
        ]

        for category in categories:
            tab = QWidget()
            tab_layout = QVBoxLayout()

            table = QTableWidget()
            table.setColumnCount(6)
            table.setHorizontalHeaderLabels([
                "Name", "Truancy Hours", "Status",
                "Meeting Date", "Meeting Time", "Meeting Location"
            ])
            # start with zero rows
            table.setRowCount(0)

            tab_layout.addWidget(table)
            tab.setLayout(tab_layout)
            self.report_tabs.addTab(tab, category)

        layout.addWidget(self.report_tabs)
        self.reports_tab.setLayout(layout)

    #Load the PDF Truancy report and extract data
    def ingest_pdf(self):
        #Open dialog to select the PDF
        file_path, _ = QFileDialog.getOpenFileName(
            self, "Open PDF File", "", "PDF Files (*.pdf)"
        )
        if not file_path:
            return
        #Parse PDF and extract student data
        self.pdf_students = extract_students_from_pdf(file_path)
        self.display_pdf_students()

    def create_new_export_file(self):
        today_str = datetime.now().strftime("%Y.%m.%d")
        base_name = f"Truancy_{today_str}"
        i = 1

        # Increment filename if needed
        while True:
            file_path = f"{base_name}_{i}.xlsx"
            if not os.path.exists(file_path):
                break
            i += 1

        # create workbook with 1 dummy sheet
        wb = Workbook()
        wb.save(file_path)

        # load it back (ensures future writes are saved correctly)
        wb = load_workbook(file_path)
        self.loaded_workbook = wb
        self.loaded_file_path = file_path

        # remove default sheet (if exists)
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # write Week 1 if PDF is already loaded
        if self.pdf_students:
            #print(f"Writing Week 1 using currently loaded PDF data...")
            sheet_name = writeWeek(wb, self.pdf_students)

            #If this is not week 1, write summary
            if not self.current_data.empty:
                from excel_compare import compare_students, write_summary_sheet
                previous_students = [
                    Student.create_empty().from_dict(d) for d in self.current_data.to_dict(orient="records")
                ]
                comparison = compare_students(self.pdf_students, previous_students)
                write_summary_sheet(wb, comparison)

            #always reorder before saving
            reorder_tabs(wb)

            wb.save(file_path)
            #print(f"Exported {sheet_name} + Summary to: {file_path}")
            #self.warning_label.setText(f"Exported {sheet_name} to: {file_path}")

        # UI feedback
        self.export_excel_button.setEnabled(True)
        self.current_file_label.setText(f"📁 New File: {os.path.basename(file_path)}")
        #print(f"New export file initialized: {file_path}")

    def display_pdf_students(self):
        self.pdf_table.clear()
        self.pdf_table.setColumnCount(9)
        self.pdf_table.setRowCount(len(self.pdf_students))
        self.pdf_table.setHorizontalHeaderLabels([
            "First Name", "Last Name", "ID", "Age", "Grade",
            "Excused", "Unexcused", "Medical", "Absence Total"
        ])
        #Pop. each row with the students data
        for row, stu in enumerate(self.pdf_students):
            self.pdf_table.setItem(row, 0, QTableWidgetItem(stu.firstName))
            self.pdf_table.setItem(row, 1, QTableWidgetItem(stu.lastName))
            self.pdf_table.setItem(row, 2, QTableWidgetItem(stu.id))
            self.pdf_table.setItem(row, 3, QTableWidgetItem(stu.age))
            self.pdf_table.setItem(row, 4, QTableWidgetItem(stu.grade))
            self.pdf_table.setItem(row, 5, QTableWidgetItem(stu.excused))
            self.pdf_table.setItem(row, 6, QTableWidgetItem(stu.unexcused))
            self.pdf_table.setItem(row, 7, QTableWidgetItem(stu.medical))
            self.pdf_table.setItem(row, 8, QTableWidgetItem(stu.absenceTotal))

    # export PDF data to text file for debugging/testing
    def export_pdf_data_to_text(self):
        if not self.pdf_students:
            print("No PDF data to export!")
            return

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Student data to text",
            "",
            "Text Files (*.txt);;All Files (*)"
        )
        if not file_path:
            return

        with open(file_path, "w", encoding="utf-8") as f:
            for student in self.pdf_students:
                f.write(self.format_student_data(student) + "\n")

        print(f"Student data written to {file_path}.")

    def format_student_data(self, student):
        return (
            f"ID: {student.id}, "
            f"First Name: {student.firstName}, "
            f"Last Name: {student.lastName}, "
            f"Age: {student.age}, "
            f"Grade: {student.grade}, "
            f"Excused: {student.excused}, "
            f"Unexcused: {student.unexcused}, "
            f"Medical: {student.medical}, "
            f"Suspension: {student.suspension}, "
            f"School Total: {student.schoolTotal}, "
            f"Attending Total: {student.attendingTotal}, "
            f"Absence Total: {student.absenceTotal}"
        )

    # placeholder for future meeting letter features
    def save_meeting_details(self, student, date_input, time_input, location_input, dialog):
        self.create_letter(student)

    def create_letter(self, student):
        doc = Document()
        doc.save()

    # export PDF data to Excel using writeWeek()
    def export_to_excel(self):
        try:
            #ensure PDF loaded before trying to export
            if not self.pdf_students:
                print("No PDF data to export!")
                return

            #if workbook is loaded, continue adding weeks there
            if self.loaded_workbook and self.loaded_file_path:
                wb = self.loaded_workbook
                file_path = self.loaded_file_path
            else:
                #if not, create a new file (starting from Week 1)
                today_str = datetime.now().strftime("%Y.%m.%d")
                file_path = f"Truancy_{today_str}.xlsx"
                wb = Workbook()
                wb.remove(wb.active)

            sheet_name = writeWeek(wb, self.pdf_students)
            reorder_tabs(wb)
            wb.save(file_path)

            #print(f"Exported to {file_path} (added tab: {sheet_name})")
            self.warning_label.setText(f"Export complete! Added {sheet_name}")

            # update stored file so next export continues in same file
            self.loaded_workbook = wb
            self.loaded_file_path = file_path
            
            #list students who have changed state to be acknowledged
            reader = ExcelReader(file_path)
            reader.read()
            self.w=acknowledgePopup(reader.fixed)
            self.w.exec()
            acknowledge(wb, self.w.ackList)
            wb.save(file_path)
            	
            

            # Compare and create summary if current_data exists
            if not self.current_data.empty:
                print("📊 Starting comparison for summary sheet...")

                previous_students = []
                for _, row in self.current_data.iterrows():
                    s = Student.create_empty()
                    s.id = row.get("id", "")
                    s.firstName = row.get("firstName", "")
                    s.lastName = row.get("lastName", "")
                    s.age = row.get("age", "")
                    s.grade = row.get("grade", "")
                    s.excused = row.get("excused", 0)
                    s.unexcused = row.get("unexcused", 0)
                    s.medical = row.get("medical", 0)
                    s.suspension = row.get("suspension", 0)
                    s.absenceTotal = row.get("absenceTotal", 0)
                    previous_students.append(s)
                #print("Old data (previous_students):")
                for s in previous_students:
                    print(s.id, s.firstName, s.lastName, "Unexcused:", s.unexcused, "Medical:", s.medical)

                #print("New data (current_students):")
                for s in self.pdf_students:
                    print(s.id, s.firstName, s.lastName, "Unexcused:", s.unexcused, "Medical:", s.medical)
                #perform data comparisons and write summary sheet. 
                comparison = compare_students(current_students=self.pdf_students, previous_students=previous_students)
                #print(f"✅ Comparison complete. Added: {len(comparison['added'])}, Removed: {len(comparison['removed'])}, Changed: {len(comparison['changed'])}")
                if "Summary" in wb.sheetnames:
                    wb.remove(wb["Summary"])
                    #print("Old Summary tab removed.")
                write_summary_sheet(wb, comparison)
                wb.save(file_path)
                #print("*New Summary tab written.")

        except Exception as e:
            #show error is export fails
            traceback.print_exc()
            msg = QMessageBox()
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setText("Excel export failed.")
            msg.setInformativeText(str(e))
            msg.setWindowTitle("Error")
            msg.exec()

class acknowledgePopup(QDialog):
	def __init__(self, inStudents):
		QDialog.__init__(self)
		self.setWindowTitle("Unacknowledged Changes")
		self.setGeometry(100, 100, 1000, 700)
		self.students = inStudents
		
		self.initUI()
	
	def initUI(self):
        	self.tabs = QTabWidget(self)
        	#self.setCentralWidget(self.tabs)

        	self.truant_tab = QWidget()
        	self.untruant_tab = QWidget()
        	self.removed_tab = QWidget()
        	#self.createTable("NOT_TRUANT")
        	#tabs for the GUI
        	self.tabs.addTab(self.createTable("TRUANT_NOT_ACKNOWLEDGED"), "Newly Truant")
        	self.tabs.addTab(self.createTable("UNTRUANT_NOT_ACKNOWLEDGED"), "Newly UnTruant")
        	self.tabs.addTab(self.createTable("REMOVAL_NOT_ACKNOWLEDGED"), "Newly Removed")
        	
        	self.tabs.resize(1000, 700)
        	
# createTable and handleItemClicked based on code from https://stackoverflow.com/questions/12366521/pyqt-checkbox-in-qtablewidget
	def createTable(self, state):
	
		tableStudents = []
		for student in self.students:
			if student.state == state:
				tableStudents.append(student)
		
		rows = len(tableStudents)
		columns = 4
		newTable = QTableWidget(rows, columns, self)
		newTable.setHorizontalHeaderLabels(["ID", "First Name", "Last Name", "Unexcused Absences"])
		for row in range(rows):
			item = QTableWidgetItem(str(tableStudents[row].id))
			item.setFlags(QtCore.Qt.ItemFlag.ItemIsUserCheckable | QtCore.Qt.ItemFlag.ItemIsEnabled)
			item.setCheckState(QtCore.Qt.CheckState.Unchecked)
			newTable.setItem(row, 0, item)
			newTable.setItem(row, 1, QTableWidgetItem(tableStudents[row].firstName))
			newTable.setItem(row, 2, QTableWidgetItem(tableStudents[row].lastName))
			newTable.setItem(row, 4, QTableWidgetItem(str(tableStudents[row].unexcused)))
		newTable.itemClicked.connect(self.handleItemClicked)
		layout = QVBoxLayout(self)
		layout.addWidget(newTable)
		self.ackList = []
		return newTable

	def handleItemClicked(self, item):
		if item.text() not in self.ackList and item.text().isdigit():
		    item.setCheckState(QtCore.Qt.CheckState.Checked)
		    print('"%s" Added' % item.text())
		    self.ackList.append(item.text())
		    print(self.ackList)
		elif item.text() in self.ackList:
		    item.setCheckState(QtCore.Qt.CheckState.Unchecked)
		    self.ackList.remove(item.text())
		    print('"%s" Removed' % item.text())
	

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = TruancyApp()
    window.show()
    sys.exit(app.exec())
