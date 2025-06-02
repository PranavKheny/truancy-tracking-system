from openpyxl import Workbook
import openpyxl
import constructor

class ExcelReader:

    #initialize reader with opening workbook and corresponding sheet name
    # fixed contains list of Students
    # ids contains list of student IDs
    def __init__(self, wb_name, sheet_name=None):
        self.wb = openpyxl.load_workbook(wb_name)

        if sheet_name and sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            # auto select most recent Week tab
            week_sheets = [name for name in self.wb.sheetnames if name.startswith("Week")]

            if not week_sheets:
                raise ValueError("No Week sheets found in workbook.")
            #grab week number from sheet names
            def week_number(name):
                try:
                    return int(name.split(" ")[1])
                except:
                    return 0
            #pick the most recent/latest week
            latest_week = sorted(week_sheets, key=week_number)[-1]
            self.ws = self.wb[latest_week]
            #print(f"Auto-selected sheet: {latest_week}")
        #store the parsed list of student objects
        self.fixed = []
        self.ids = []


    # function to read in values from current opened workbook
    def read(self):
        # get labels at top of workbook
        labels = self._readLabels()
        print("Headers found in Excel:", labels)

        # get list of dictionaries containing student info from sheet
        #extract student info rows, map to dictionaries
        mapped_students = self._readStudents(labels)

        # to help debug and preview the first parsed student dictionary
        '''
        if mapped_students:
            print("sample row:", mapped_students[0])
        print(f"parsed row count: {len(mapped_students)}")
        '''
        #convert each dictionary to a student obj.
        for student in mapped_students:
            student_obj = constructor.Student(
                student.get("id", ""),
                str(student.get("First Name", "") or "").strip(),
                str(student.get("Last Name", "") or "").strip(),
                student.get("Age", ""),
                student.get("Grade", ""),
                float(student.get("Excused Absences") or 0.0),
                float(student.get("Unexcused Absences") or 0.0),
                float(student.get("Medically Excused") or 0.0),
                float(student.get("Suspension Hours") or 0.0),
                0.0,  # schoolTotal
                0.0,  # attendingTotal
                float(student.get("Total Absences (minus suspension hours)") or 0.0)

            )
            self.fixed.append(student_obj)



    #helper function to read labels (row 1) in opened workbook to get headers
    def _readLabels(self):
        labels = []
        row = self.ws[1]
        for cell in row:
            labels.append(cell.value)
        print("Headers found in Excel:", labels)
        return labels

    #helper function to read all student entries in workbook
    #creates a temporary list of dictionaries to be used in student constructors
    def _readStudents(self, labels):
        temp_list = []
        for row in self.ws.iter_rows(min_row=2):
            student = {}
            i = 0
            #for each cell, save to dictionary based on labels obtained in _readLabels

            for cell in row:
                val = cell.value
                label = labels[i]
                #if the label is ID, save the id and also change its type to get rid of .0 at end
                if label == "id":
                    if val is None:
                        break  # skip blank rows
                    string_val = str(val)
                    student[label] = string_val.split(".")[0]
                    self.ids.append(student[label])
                else:
                    if label in ["Last Name", "First Name"]:
                        student[label] = str(val or "").strip()
                    else:
                        student[label] = val
                i += 1
            if student.get("id"):  #only add valid rows with ids
                temp_list.append(student)
        return temp_list

    #console testing
    '''
    def print(self):
        constructor.Student.printHeaders()
        for child in self.fixed:
            child.print()
    '''

#For testing
if __name__ == "__main__":
    # For testing only
    e = ExcelReader("SimpleTest1.xlsx", "Week 1")
    e.read()
    e.print()
