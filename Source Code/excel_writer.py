from openpyxl import Workbook
import openpyxl
from datetime import date
import constructor

#writes the headers to the first row of uninitialized spreadsheet "ws"
def writeHeaders(ws):
	headers = ["id", "Last Name", "First Name", "Age", "Grade", "Custodian", "Address", "Suspension Hours", "Excused Absences", "UNEXCUSED_UNINITIALIZED", "Total Absences\n(minus suspension hours)", "Outcome of Correspondence", "Status"]
	for i in range(len(headers)):
		ws.cell(1, i+1, headers[i])
		ws.column_dimensions[chr(i+65)].width = 20
	ws.row_dimensions[1].height = 35

#writes data from student object array to a new tab called 'Week X'
def writeWeek(wb, students, date=date.today()):
    #remove any default sheets from workbook creation
    for name in wb.sheetnames:
        if name.lower() in ["sheet", "sheet1", "summary"]:
            ws_to_remove = wb[name]
            wb.remove(ws_to_remove)
            break
    # get all existing sheet names that match format "Week X"
    week_nums = []
    for name in wb.sheetnames:
        if name.startswith("Week "):
            try:
                week_num = int(name.replace("Week ", ""))
                week_nums.append(week_num)
            except:
                pass

    # if no previous week found, this is week 1
    next_week = max(week_nums) + 1 if week_nums else 1
    sheet_name = f"Week {next_week}"

    # if there is already data, set OLDws to newest week sheet
    if len(wb.sheetnames)>0:
    	OLDws = wb.worksheets[0]
    	
    # create new sheet
    ws = wb.create_sheet(title=sheet_name, index=1)  # inserts as leftmost tab

    # copy previous week's data to new sheet
    if len(wb.sheetnames)>1:
    	copier = openpyxl.worksheet.copier.WorksheetCopy(OLDws, ws)
    	copier.copy_worksheet()
    
    #if len(wb.sheetnames)>1:
    #	wb.copy_worksheet(wb.worksheets[wb.index(ws)-1], ws)
    
    # if the headers are not populated, populate them
    if (ws.cell(1, 1).value != "id"):
    	writeHeaders(ws)

    #defines to which static column the current weeks unexcused absences should be written (the column that is overwritten every week)
    standardUnexAbsCol = 10
	
    #compute weekCol to be the column the dynamic column at the right to which new data should be added
    weekCol=standardUnexAbsCol
    #move through columns until the header is "Outcome of Correspondence", new week data goes in column before this
    while ws.cell(1, weekCol).value != "Outcome of Correspondence":
    	weekCol+=1
	
    #insert new column and update headers for dynamic and static week columns
    ws.insert_cols(weekCol)
    ws.column_dimensions[chr(weekCol+65)].width = 30
    ws.column_dimensions[chr(weekCol+64)].width = 25
    ws.cell(1, standardUnexAbsCol, date.strftime('%m/%d/%y') + "\nUnexcused Absences")
    ws.cell(1, weekCol, date.strftime('%m/%d/%y') + "\nUnexcused Absences")
    i=2
	
    #background colors for cells
    noBackground = openpyxl.styles.PatternFill(fill_type=None)
    redBackground = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

	
    #erase the static unexAbsCol; if a student dropped off new report their current cell should be blank
    while ws.cell(i,1).value != None:
    	ws.cell(i,standardUnexAbsCol,"")
    	#erase any background color, if present
    	ws.cell(i,standardUnexAbsCol).fill = noBackground
    	i+=1
	
    #for every child in the given array, find the row with their student id and update the data in that row
    #if the student id not present, put the data in a new row at the bottom
    for child in students:
    	i=2
    	while ws.cell(i,1).value != child.id and ws.cell(i,1).value != None:
    		i+=1
    	
    	ws.cell(i,1,child.id)
    	ws.cell(i,2,child.lastName)
    	ws.cell(i,3,child.firstName)
    	ws.cell(i,4,child.age)
    	ws.cell(i,5,child.grade)
    	ws.cell(i,8,child.suspension)
    	ws.cell(i,9,child.excused)
    	ws.cell(i,10,child.unexcused)
    	ws.cell(i,11,child.schoolTotal)
    	ws.cell(i,weekCol,child.unexcused)
    	if ws.cell(i, weekCol+2).value == None:
    		ws.cell(i,weekCol+2, 'NOT_TRUANT')
		
    	#mark students who have reached the threshold in red, update state
    	if float(child.unexcused) > constructor.Student.redThreshold:
    		ws.cell(i,weekCol).fill = redBackground
    		ws.cell(i,standardUnexAbsCol).fill = redBackground
    		#TODO: replace with enumerable
    		if ws.cell(i, weekCol+2).value == 'NOT_TRUANT' or ws.cell(i, weekCol+2).value == 'REMOVAL_NOT_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'REMOVAL_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'UNTRUANT_NOT_ACKNOWLEDGED':
    			ws.cell(i,weekCol+2, 'TRUANT_NOT_ACKNOWLEDGED')
    	elif ws.cell(i, weekCol+2).value == 'TRUANT_NOT_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'TRUANT_ACKNOWLEDGED':
    		ws.cell(i,weekCol+2, 'UNTRUANT_NOT_ACKNOWLEDGED')
	
    i=2
    while ws.cell(i,1).value != None:
    	if ws.cell(i, standardUnexAbsCol).value == "" and ws.cell(i, weekCol+2).value != 'REMOVAL_ACKNOWLEDGED':
    		ws.cell(i, weekCol+2).value = 'REMOVAL_NOT_ACKNOWLEDGED'
    	i += 1

    return sheet_name  # return name for tracking

# reorders the sheets to ensure Summary is always first, and weeks in reverse order (newest left)
def reorder_tabs(wb):
    summary = None
    weeks = []
    #find and categorize tabs
    for sheet in wb.worksheets:
        if sheet.title.lower() == "summary":
            summary = sheet
        elif sheet.title.startswith("Week "):
            try:
                week_num = int(sheet.title.replace("Week ", ""))
                weeks.append((week_num, sheet))
            except:
                continue

    #sort weeks in reverse so newest is always first
    weeks.sort(reverse=True, key=lambda x: x[0])
    #build new sheet order - if summary exist -first, then sorted weeks
    new_order = [summary] if summary else []
    new_order += [ws for _, ws in weeks]
    #apply the new tab order
    wb._sheets = new_order  # reorder internal sheet list
    
# updates the given student ids to be in their corresponding "acknowledged state"
def acknowledge(wb, students):

	# open the sheet for the most recent week
	ws = wb.worksheets[0]
	if ws.title == "Summary":
		ws = wb.worksheets[1]
	
	#compute weekCol to be the column the dynamic column at the right to which new data should be added
	weekCol=10
	#move through columns until the header is "Outcome of Correspondence", new week data goes in column before this
	while ws.cell(1, weekCol).value != "Outcome of Correspondence":
		weekCol+=1
	
	#for each student who has been acknowledged, find their row and update their state to be acknowledged
	for child in students:
		i=2
		while str(ws.cell(i,1).value) != str(child) and ws.cell(i,1).value != None:
			i+=1
		if ws.cell(i, weekCol+1).value == 'TRUANT_NOT_ACKNOWLEDGED':
			ws.cell(i, weekCol+1, "TRUANT_ACKNOWLEDGED")
		elif ws.cell(i, weekCol+1).value == 'UNTRUANT_NOT_ACKNOWLEDGED':
			ws.cell(i, weekCol+1, "UNTRUANT_ACKNOWLEDGED")
		else:
			ws.cell(i, weekCol+1, "REMOVAL_ACKNOWLEDGED")
    		
		
	

