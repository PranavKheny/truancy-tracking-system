from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# compare current (PDF) vs previous student data (Excel uploaded)
def compare_students(current_students, previous_students):
    # float conversion to handle invalid
    def safe_float(val):
        try:
            return float(val)
        except:
            return 0.0
    #dict to hold student IDs for quick lookups
    current_ids = {s.id: s for s in current_students}
    previous_ids = {s.id: s for s in previous_students}

    #results
    added = [] #in current, but not prev
    removed = [] #in prev, but not current
    changed = [] #in both, but a valid has changed

    #iterate through the current students
    for id, student in current_ids.items():
        #if new ID, append to added
        if id not in previous_ids:
            added.append(student)
        else:
            #compare with previous record
            prev = previous_ids[id]
            diffs = {}
            #calculate differences
            unexcused_diff = safe_float(student.unexcused) - safe_float(prev.unexcused)
            medical_diff = safe_float(student.medical) - safe_float(prev.medical)
            suspension_diff = safe_float(student.suspension) - safe_float(prev.suspension)
            excused_diff = safe_float(student.excused) - safe_float(prev.excused)
            misc_diff = 0  # Reserved for future
            #store rounded variables
            diffs["unexcused"] = round(unexcused_diff, 2)
            diffs["medical"] = round(medical_diff, 2)
            diffs["suspension"] = round(suspension_diff, 2)
            diffs["excused"] = round(excused_diff, 2)
            diffs["misc"] = round(misc_diff, 2)

            # Only add to changed list if theres non-zero differences
            if any(val != 0 for val in diffs.values()):
                changed.append((student, diffs))
    #if not in current, but in prev, append to removed
    for id, student in previous_ids.items():
        if id not in current_ids:
            removed.append(student)

    return {
        "added": added,
        "removed": removed,
        "changed": changed
    }


def write_summary_sheet(wb, comparison):
    #Remove old summary sheet
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    #make a summary sheet in index 0 - Always should be first when opening excel file
    ws = wb.create_sheet("Summary", index=0)
    wb.active = 0
    row = 1

    #helper function for the 3 sections
    def write_section(title, students, color_hex, diff=False):
        nonlocal row
        #start styling section headings
        title_font = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill("solid", fgColor=color_hex)
        #merge cells for headers
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        title_cell = ws.cell(row=row, column=1, value=title)
        title_cell.font = title_font
        title_cell.fill = title_fill
        row += 1
        #write the headers to display on summary sheet
        headers = [
            "Student", 
            "Unexcused Absences", 
            "Medically Excused", 
            "Suspension Hours", 
            "Excused Absences", 
            "Misc. Hours"
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = title_fill
        row += 1
        #write out data for students
        for s in students:
            if diff:
                #if diff is true, each item has a pair (student obj, difference dict)
                #unpack it into base - student object and diff - the changes
                base, diffs = s

                # Skip if all differences are zero
                if all(
                    round(diffs.get(field, 0) or 0, 2) == 0
                    for field in ["unexcused", "medical", "suspension", "excused", "misc"]
                ):
                    continue
                #name formatting
                name = f"{base.lastName or ''}, {base.firstName or ''}".strip(", ").strip()
                if not name:
                    name = f"(ID: {base.id})"
                ws.cell(row=row, column=1, value=name)
                #write changes with delta formatting
                #helper method since we will be writing the values multiple times for each cat
                def write_change(col, val, delta):
                    txt = f"{val} ({'+' if delta > 0 else ''}{delta})" if delta != 0 else f"{val} (+0)"
                    cell = ws.cell(row=row, column=col, value=txt)
                    if delta > 0:
                        cell.font = Font(color="008000")
                    elif delta < 0:
                        cell.font = Font(color="FF0000")
                    else:
                        cell.font = Font(color="666666")

                write_change(2, base.unexcused or 0, diffs.get("unexcused", 0))
                write_change(3, base.medical or 0, diffs.get("medical", 0))
                write_change(4, base.suspension or 0, diffs.get("suspension", 0))
                write_change(5, base.excused or 0, diffs.get("excused", 0))
                write_change(6, 0, diffs.get("misc", 0))

            else:
                #if added/removed - write student's basic infor
                first = (s.firstName or "").strip()
                last = (s.lastName or "").strip()
                name = f"{last}, {first}".strip(", ") if (first or last) else f"(ID: {s.id})"
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=s.unexcused or 0)
                ws.cell(row=row, column=3, value=s.medical or 0)
                ws.cell(row=row, column=4, value=s.suspension or 0)
                ws.cell(row=row, column=5, value=s.excused or 0)
                ws.cell(row=row, column=6, value=0)

            row += 1
        #space between sections
        row += 1
    #if no changes at all, return a message on the summary sheet instead
    if not (comparison["added"] or comparison["removed"] or comparison["changed"]):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value="No changes between current and previous data")
        cell.font = Font(bold=True, color="008000")
        return
    #the sections of the summary calling write section
    write_section("Additions", comparison["added"], "2F75B5")
    write_section("Changes", comparison["changed"], "548235", diff=True)
    write_section("Removed", comparison["removed"], "7030A0")
    #resize col widths to be easier to read
    for col in range(1, 7):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 25
