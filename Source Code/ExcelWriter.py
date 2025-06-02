'''
	This file is used to write student objects to a new or existing Excel file. If a new Excel file is being used, writeHeaders should be called before any calls to writeWeek to ensure the file is properly initialized. Once initialized, writeWeek takes an array of student data objects and writes them to an Excel file, respecting the data already in the file, if any, and treating the current data as a new week.
'''

import openpyxl
import constructor
from datetime import date

#writes the headers to the first row of uninitialized spreadsheet "ws"
def writeHeaders(ws):
	headers = ["id", "Last Name", "First Name", "Age", "Grade", "Custodian", "Address", "Suspension Hours", "Excused Absences", "UNEXCUSED_UNINITIALIZED", "Total Absences\n(minus suspension hours)", "Outcome of Correspondence", "STATE"]
	for i in range(len(headers)):
		ws.cell(1, i+1, headers[i])
		ws.column_dimensions[chr(i+65)].width = 20
	ws.row_dimensions[1].height = 35

#writes the data from the array of student objects "students" to the spreadsheet "ws"
def writeWeek(ws, students, date=date.today()):

	#defines to which static column the current weeks unexcused absences should be written (the column that is overwritten every week)
	standardUnexAbsCol = 10
	
	#compute weekCol to be the column the dynamic column at the right to which new data should be added
	weekCol=standardUnexAbsCol
	#move through columns until the header is "Outcome of Correspondence", new week data goes in column before this
	while ws.cell(1, weekCol).value != "Outcome of Correspondence":
		weekCol+=1
	
	#insert new column and update headers for dynamic and static week columns
	ws.insert_cols(weekCol)
	ws.column_dimensions[chr(weekCol+65)].width = 30
	ws.column_dimensions[chr(weekCol+64)].width = 25
	ws.cell(1, standardUnexAbsCol, date.strftime('%m/%d/%y') + "\nUnexcused Absences")
	ws.cell(1, weekCol, date.strftime('%m/%d/%y') + "\nUnexcused Absences")
	i=2
	
	#background colors for cells
	noBackground = openpyxl.styles.PatternFill(fill_type=None)
	redBackground = openpyxl.styles.PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

	
	#erase the static unexAbsCol; if a student dropped off new report their current cell should be blank
	while ws.cell(i,1).value != None:
		ws.cell(i,standardUnexAbsCol,"")
		#erase any background color, if present
		ws.cell(i,standardUnexAbsCol).fill = noBackground
		i+=1
	
	#for every child in the given array, find the row with their student id and update the data in that row
	#if the student id not present, put the data in a new row at the bottom
	i=2
	for child in students:
		while ws.cell(i,1).value != child.id and ws.cell(i,1).value != None:
			i+=1
		
		ws.cell(i,1,child.id)
		ws.cell(i,2,child.lastName)
		ws.cell(i,3,child.firstName)
		ws.cell(i,4,child.age)
		ws.cell(i,5,child.grade)
		ws.cell(i,8,child.suspension)
		ws.cell(i,9,child.excused)
		ws.cell(i,10,child.unexcused)
		ws.cell(i,11,child.schoolTotal)
		ws.cell(i,weekCol,child.unexcused)
		if ws.cell(i, weekCol+2).value == None:
			ws.cell(i,weekCol+2, 'NOT_TRUANT')
		
		#mark students who have reached the threshold in red, update state
		if child.unexcused > constructor.Student.redThreshold:
			ws.cell(i,weekCol).fill = redBackground
			ws.cell(i,standardUnexAbsCol).fill = redBackground
			#TODO: replace with enumerable
			if ws.cell(i, weekCol+2).value == 'NOT_TRUANT' or ws.cell(i, weekCol+2).value == 'REMOVAL_NOT_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'REMOVAL_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'UNTRUANT_NOT_ACKNOWLEDGED':
				ws.cell(i,weekCol+2, 'TRUANT_NOT_ACKNOWLEDGED')
		elif ws.cell(i, weekCol+2).value == 'TRUANT_NOT_ACKNOWLEDGED' or ws.cell(i, weekCol+2).value == 'TRUANT_ACKNOWLEDGED':
			ws.cell(i,weekCol+2, 'UNTRUANT_NOT_ACKNOWLEDGED')
	
	i=2
	while ws.cell(i,1).value != None:
		if ws.cell(i, standardUnexAbsCol).value == "" and ws.cell(i, weekCol+2).value != 'REMOVAL_ACKNOWLEDGED':
			ws.cell(i, weekCol+2).value = 'REMOVAL_NOT_ACKNOWLEDGED'
		i += 1
	
		
