from openpyxl import Workbook
import openpyxl
import constructor

class ExcelReader:
    
    #initialize reader with opening workbook and corresponding sheet name
    # fixed contains list of Students
    # ids contains list of student IDs
    def __init__(self, wb_name, sheet_name=None):
        self.wb = openpyxl.load_workbook(wb_name)

        if sheet_name and sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            # auto select most recent Week tab
            week_sheets = [name for name in self.wb.sheetnames if name.startswith("Week")]

            if not week_sheets:
                raise ValueError("No Week sheets found in workbook.")
            #grab week number from sheet names
            def week_number(name):
                try:
                    return int(name.split(" ")[1])
                except:
                    return 0
            #pick the most recent/latest week
            latest_week = sorted(week_sheets, key=week_number)[-1]
            self.ws = self.wb[latest_week]
            #print(f"Auto-selected sheet: {latest_week}")
        #store the parsed list of student objects
        self.fixed = []
        self.ids = []
        
        
    # function to read in values from current opened workbook
    def read(self):
        #get labels at top of workbook
        labels = self._readLabels()
        #get current week value
        date = self._getDate(labels)
        #get list of dictionaries containing student info from sheet
        mapped_students = self._readStudents(labels)

        #for each student in the dictionary, create a Student object and push to self.fixed
        for student in mapped_students:
            student_obj = constructor.Student(
                student["id"],
                student["First Name"],
                student["Last Name"],
                student["Age"],
                student["Grade"],
                student["Excused Absences"],
                student[date],
                0.0,
                student["Suspension Hours"],
                0.0,
                0.0,
                student[f"Total Absences\n(minus suspension hours)"],
                student["Status"]
                )
            self.fixed.append(student_obj)

        
    #helper function to read labels (row 1) in opened workbook
    def _readLabels(self):
        labels = []
        row = self.ws[1]
        for cell in row:
            labels.append(cell.value)
        return labels
    
    #helper function to get the current week we are in for spreadsheet
    def _getDate(self, labels):
        return labels[9]

    #helper function to read all student entries in workbook
    #creates a temporary list of dictionaries to be used in student constructors
    def _readStudents(self, labels):
        temp_list = []
        #iterate through each row
        for row in self.ws.iter_rows(min_row=2):
            student = {}
            i = 0
            #for each cell, save to dictionary based on labels obtained in _readLabels
            for cell in row:
                val = cell.value
                #if the label is ID, save the id and also change its type to get rid of .0 at end
                if(labels[i] == "id"):
                    string_val = str(val)
                    student[labels[i]] = string_val.split(".")[0]
                    self.ids.append(student[labels[i]])
                else:
                    student[labels[i]] = val
                i += 1
            temp_list.append(student)
        
        return temp_list
        
    def print(self):
        constructor.Student.printHeaders()
        for child in self.fixed:
            child.print()


